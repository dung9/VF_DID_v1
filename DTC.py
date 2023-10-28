from tkinter import *
import tkinter
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from tkinter.ttk import *
import can
import subprocess
import tkinter as tk

from openpyxl import load_workbook

DTC = Tk()
DTC.title("New Window")
DTC.geometry("1200x600") 
DTC.resizable(0, 0)

ALL_Status = IntVar()
VCU_Status = IntVar()
DCDC_Status = IntVar()
POD_Status = IntVar()
OBC_Status = IntVar()
EDS_Status= IntVar()
BMS_Status= IntVar()
GS_Status= IntVar()
IDB_Status= IntVar()
RCU_Status= IntVar()
EPS_Status= IntVar()
ACM_Status= IntVar()
BCM_Status= IntVar()
BCM_BPM_Status= IntVar()
CCU1_Status= IntVar()
XGW_Status= IntVar()
APM_Status= IntVar()
SHVU_F_Status= IntVar()
SHVU_R_Status= IntVar()
OCS_Status= IntVar()
MHU_Status= IntVar()
HUD_Status= IntVar()
AVAS_Status= IntVar()
AP_ECU_Status= IntVar()
FCAM_Status= IntVar()
MCR_FL_RADAR_Status= IntVar()
MCR_FR_RADAR_Status= IntVar()
MCR_RR_RADAR_Status= IntVar()
MCR_RL_RADAR_Status= IntVar()
MFR1_RADAR_Status= IntVar()

ECU_Frame = tk.LabelFrame(DTC,bg = '#c3c3c3',text= 'ECU',borderwidth=1,relief='solid')
ECU_Frame.pack(side=tk.LEFT)
ECU_Frame.pack_propagate(False)
ECU_Frame.configure(width=100,height=900)

Excel_frame = tk.LabelFrame(DTC,bg = '#c3c3c3',text= 'ECU DTC',borderwidth=1,relief='solid')
Excel_frame.pack(side=tk.TOP,padx=1,pady=1)
Excel_frame.pack_propagate(False)
Excel_frame.configure(width=1400,height=900)

Header = ["No","DTC","Description","Staus"]
style = ttk.Style()
style.theme_use('clam')
tree = ttk.Treeview(Excel_frame, column=("c1", "c2","c3","c4"), show='headings')
tree.pack(expand=True, fill='x')

count = 0
for i in range(4):
    count += 1
    tree.column("# "+ str(count), anchor=CENTER,width=300)
    tree.heading("# "+ str(count), text=Header[i])
    tree.pack(expand=True, fill='y')

scrollbar = ScrolledText(ECU_Frame)
scrollbar.pack( side = RIGHT, fill=Y )

ALL_Status.set(0)
VCU_Status.set(0)
DCDC_Status.set(0)
POD_Status.set(0)
OBC_Status.set(0)
EDS_Status.set(0)
BMS_Status.set(0)
GS_Status.set(0)
IDB_Status.set(0)
RCU_Status.set(0)
EPS_Status.set(0)
ACM_Status.set(0)
BCM_Status.set(0)
BCM_BPM_Status.set(0)
CCU1_Status.set(0)
XGW_Status.set(0)
APM_Status.set(0)
SHVU_F_Status.set(0)
SHVU_R_Status.set(0)
OCS_Status.set(0)
MHU_Status.set(0)
HUD_Status.set(0)
AVAS_Status.set(0)
AP_ECU_Status.set(0)
FCAM_Status.set(0)
MCR_FL_RADAR_Status.set(0)
MCR_FR_RADAR_Status.set(0)
MCR_RR_RADAR_Status.set(0)
MCR_RL_RADAR_Status.set(0)
MFR1_RADAR_Status.set(0)  

ALL = tk.Checkbutton(scrollbar, text='ALL', bg='white', anchor='w',variable= ALL_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=ALL)
scrollbar.insert('end', '\n')
VCU = tk.Checkbutton(scrollbar, text='VCU', bg='white', anchor='w',variable= VCU_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=VCU)
scrollbar.insert('end', '\n')
DCDC = tk.Checkbutton(scrollbar, text='DCDC', bg='white', anchor='w',variable= DCDC_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=DCDC)
scrollbar.insert('end', '\n')
POD = tk.Checkbutton(scrollbar, text='POD', bg='white', anchor='w',variable= POD_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=POD)
scrollbar.insert('end', '\n')
OBC = tk.Checkbutton(scrollbar, text='OBC', bg='white', anchor='w',variable= OBC_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=OBC)
scrollbar.insert('end', '\n')
EDS = tk.Checkbutton(scrollbar, text='EDS', bg='white', anchor='w',variable= EDS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=EDS)
scrollbar.insert('end', '\n')
BMS = tk.Checkbutton(scrollbar, text='BMS', bg='white', anchor='w',variable= BMS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=BMS)
scrollbar.insert('end', '\n')
GS = tk.Checkbutton(scrollbar, text='GS', bg='white', anchor='w',variable= GS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=GS)
scrollbar.insert('end', '\n')
IDB = tk.Checkbutton(scrollbar, text='IDB', bg='white', anchor='w',variable= IDB_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=IDB)
scrollbar.insert('end', '\n')
RCU = tk.Checkbutton(scrollbar, text='RCU', bg='white', anchor='w',variable= RCU_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=RCU)
scrollbar.insert('end', '\n')
EPS = tk.Checkbutton(scrollbar, text='EPS', bg='white', anchor='w',variable= EPS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=EPS)
scrollbar.insert('end', '\n')
ACM = tk.Checkbutton(scrollbar, text='ACM', bg='white', anchor='w',variable= ACM_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=ACM)
scrollbar.insert('end', '\n')
BCM = tk.Checkbutton(scrollbar, text='BCM', bg='white', anchor='w',variable= BCM_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=BCM)
scrollbar.insert('end', '\n')
BPM = tk.Checkbutton(scrollbar, text='BPM', bg='white', anchor='w',variable= BCM_BPM_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=BPM)
scrollbar.insert('end', '\n')
CCU = tk.Checkbutton(scrollbar, text='CCU', bg='white', anchor='w',variable= CCU1_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=CCU)
scrollbar.insert('end', '\n')
XGW = tk.Checkbutton(scrollbar, text='XGW', bg='white', anchor='w',variable= XGW_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=XGW)
scrollbar.insert('end', '\n')
APM = tk.Checkbutton(scrollbar, text='APM', bg='white', anchor='w',variable= APM_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=APM)
scrollbar.insert('end', '\n')
SHVU_F = tk.Checkbutton(scrollbar, text='SHVU_F', bg='white', anchor='w',variable= SHVU_F_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=SHVU_F)
scrollbar.insert('end', '\n')
SHVU_R = tk.Checkbutton(scrollbar, text='SHVU_R', bg='white', anchor='w',variable= SHVU_R_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=SHVU_R)
scrollbar.insert('end', '\n')
OCS = tk.Checkbutton(scrollbar, text='OCS', bg='white', anchor='w',variable= OCS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=OCS)
scrollbar.insert('end', '\n')
MHU = tk.Checkbutton(scrollbar, text='MHU', bg='white', anchor='w',variable= MHU_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MHU)
scrollbar.insert('end', '\n')
HUD = tk.Checkbutton(scrollbar, text='HUD', bg='white', anchor='w',variable= HUD_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=HUD)
scrollbar.insert('end', '\n')
AVAS = tk.Checkbutton(scrollbar, text='AVAS', bg='white', anchor='w',variable= AVAS_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=AVAS)
scrollbar.insert('end', '\n')
AP_ECU = tk.Checkbutton(scrollbar, text='AP_ECU', bg='white', anchor='w',variable= AP_ECU_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=AP_ECU)
scrollbar.insert('end', '\n')
FCAM = tk.Checkbutton(scrollbar, text='FCAM', bg='white', anchor='w',variable= FCAM_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=FCAM)
scrollbar.insert('end', '\n')
MCR_FL = tk.Checkbutton(scrollbar, text='MCR_FL', bg='white', anchor='w',variable= MCR_FL_RADAR_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MCR_FL)
scrollbar.insert('end', '\n')
MCR_FR = tk.Checkbutton(scrollbar, text='MCR_FR', bg='white', anchor='w',variable= MCR_FR_RADAR_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MCR_FR)
scrollbar.insert('end', '\n')
MCR_RR = tk.Checkbutton(scrollbar, text='MCR_RR', bg='white', anchor='w',variable= MCR_RR_RADAR_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MCR_RR)
scrollbar.insert('end', '\n')
MCR_RL = tk.Checkbutton(scrollbar, text='MCR_RL', bg='white', anchor='w',variable= MCR_RL_RADAR_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MCR_RL)
scrollbar.insert('end', '\n')
MFR = tk.Checkbutton(scrollbar, text='MFR', bg='white', anchor='w',variable= MFR1_RADAR_Status,onvalue=1, offvalue=0)
scrollbar.window_create('end', window=MFR)
scrollbar.insert('end', '\n')    
ECU_Frame = tk.LabelFrame(DTC,bg = '#c3c3c3',text= 'ECU',borderwidth=1,relief='solid')
try:
    bus = can.interface.Bus(bustype='vector', app_name='VINFAST', channel= 1, bitrate=500000)
    print('Connect with can box is done')
except can.CanError:
    print('Check can box is connect or not')
#BCM DTC
BCM_DiagRq_DTC = can.Message(arbitration_id= 0x681,
                 is_extended_id= 0,
                 dlc= 8,data=[3, 25, 2, 9, 255, 255, 255, 255],
                 check = False)
BCM_DiagRq_DTC1 = can.Message(arbitration_id= 0x681,
                 is_extended_id= 0,
                 dlc= 8,data=[48, 0, 20, 255, 255, 255, 255, 255],
                 check = False)
#XGW DTC
XGW_DiagRq_DTC = can.Message(arbitration_id= 0x682,
                 is_extended_id= 0,
                 dlc= 8,data=[3, 25, 2, 9, 255, 255, 255, 255],
                 check = False)
XGW_DiagRq_DTC1 = can.Message(arbitration_id= 0x682,
                 is_extended_id= 0,
                 dlc= 8,data=[48, 0, 20, 255, 255, 255, 255, 255],
                 check = False)
def ECU_Req_INFOR(message):
    try:
        bus.send(message)
    except can.CanError:
        print("Msg cannot send")

def ECU_Resp(ECU_ID,message):
    global ECU_LIST
    ECU_LIST = [0 for i in range(1000)]
    global Index,DTC_Cout
    Index = 0 
    DTC_Cout = 0
    DTC_STATUS = 0
    while DTC_STATUS == 0:
        mess = bus.recv(timeout = 2)
        try:
            if mess.arbitration_id == ECU_ID:
                if mess.data[0] == 16:
                    ECU_Req_INFOR(message)
                if DTC_Cout == 0:
                    for i in range(5,8):
                        ECU_LIST[Index] = mess.data[i]
                        Index +=1
                else:
                    for i in range(1,8):    
                        ECU_LIST[Index] = mess.data[i]
                        Index +=1
                DTC_Cout +=1
        except AttributeError:
            print('Can not read DTC')
            DTC_STATUS = 1
def DTC_Read(mesage,mesage1,ID):

    DTC_List = load_workbook('BCM_DTC_VF33_VFe34s.xlsx',data_only=True)
    DTC_List_sheet = DTC_List.get_sheet_by_name('DTC List')
    
    ECU_Req_INFOR(mesage)
    ECU_Resp(ID,mesage1)      
    cout = 0
    for j in range(Index//4):
        DTC_Name = str(hex(ECU_LIST[cout])[2:].zfill(2))+ str(hex(ECU_LIST[cout+1])[2:].zfill(2))+ str(hex(ECU_LIST[cout+2])[2:].zfill(2))
        for i in range(1,DTC_List_sheet.max_row):
            Discription = "None"
            if (DTC_Name.lower() in str(DTC_List_sheet.cell(row=i,column=2).value).lower()) == True:
                Discription = DTC_List_sheet.cell(row=i,column=5).value
                break
        print("DTC" + str(j)+":" + hex(ECU_LIST[cout])[2:].zfill(2)+ hex(ECU_LIST[cout+1])[2:].zfill(2)+ hex(ECU_LIST[cout+2])[2:].zfill(2)+ hex(ECU_LIST[cout+3])[2:].zfill(2) + '\n')
        tree.insert('', 'end',iid =j,values=(j, hex(ECU_LIST[cout])[2:].zfill(2)+ hex(ECU_LIST[cout+1])[2:].zfill(2)+ hex(ECU_LIST[cout+2])[2:].zfill(2),Discription,"None"))
        tree.pack()        
        cout +=4
    print(ECU_LIST)
    print(Index)
    print(DTC_Cout)
DTC_Read(BCM_DiagRq_DTC,BCM_DiagRq_DTC1,0x601)
mainloop()