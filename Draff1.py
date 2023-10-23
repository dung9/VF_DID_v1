from openpyxl import Workbook, load_workbook

Draf = load_workbook("Daff.xlsx",data_only= True)
ReadDID_sheet = Draf.get_sheet_by_name('ECU_DID')

print(ReadDID_sheet.cell(row=2,column=3).value)