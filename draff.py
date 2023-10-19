from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

VSR_BOM = load_workbook("VF6 LS FRS VSR.xlsx")
VSR_BOM_sheet = VSR_BOM.get_sheet_by_name('FRS6.1.1.1 US')
BOM = load_workbook("rich-filter-results-231017-095647.xlsx")
BOM_sheet = BOM.get_sheet_by_name('Rich Filter Results')

for i in range(1,BOM_sheet.max_row + 1):
    if ( 'PLUS' in BOM_sheet.cell(row = i,column = 6).value) == True:
        for j in range(1,VSR_BOM_sheet.max_row + 1):
            if VSR_BOM_sheet.cell(row = j,column = 14).value == 'X':
                if BOM_sheet.cell(row = i,column = 5).value == VSR_BOM_sheet.cell(row = j,column = 17).value:
                    if (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 8).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 8).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 9).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 9).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 10).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 10).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 11).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 11).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 12).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 12).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 13).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 13).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 14).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 14).value
                    elif (VSR_BOM_sheet.cell(row = j,column = 6).value in BOM_sheet.cell(row = 1,column = 15).value) == True:
                        VSR_BOM_sheet.cell(row = j,column = 3).value = BOM_sheet.cell(row = i,column = 15).value

VSR_BOM.save("VF6 LS FRS VSR.xlsx")
