# importing only  those functions  
from tkinter import *
from tkinter import PhotoImage
import tkinter 
from tkinter.ttk import * 
import subprocess
import cantools
from tkinter import Tk, filedialog
import numpy as np
from openpyxl import load_workbook,Workbook  
from tkinter import ttk
from pathlib import Path
import tkinter as tk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from tkinter.scrolledtext import ScrolledText
import time
import can
import can.interfaces.vector
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from ReadDID import*

# creating tkinter window 
root = Tk() 
root.title('VTOOL') 
tabControl = ttk.Notebook(root)

global DID_Infor 
HW = 1
SW = 2
HW_rv = 3
bl = 4
try:
    bus = can.interface.Bus(bustype='vector', app_name='VINFAST', channel= 0, bitrate=500000)
    print('Connect with can box is done')
except can.CanError:
    print('Check can box is connect or not')

# creating variable
ALL_Status = IntVar()
VCU_Status = IntVar()
DCDC_Status = IntVar()
POD_Status = IntVar()
OBC_Status = IntVar()
EDS_Status= IntVar()
BMS_Status= IntVar()
GS_Status= IntVar()
IDB_Status= IntVar()
RCU_Status= IntVar()
EPS_Status= IntVar()
ACM_Status= IntVar()
BCM_Status= IntVar()
BCM_BPM_Status= IntVar()
CCU1_Status= IntVar()
XGW_Status= IntVar()
APM_Status= IntVar()
SHVU_F_Status= IntVar()
SHVU_R_Status= IntVar()
OCS_Status= IntVar()
MHU_Status= IntVar()
HUD_Status= IntVar()
AVAS_Status= IntVar()
AP_ECU_Status= IntVar()
FCAM_Status= IntVar()
MCR_FL_RADAR_Status= IntVar()
MCR_FR_RADAR_Status= IntVar()
MCR_RR_RADAR_Status= IntVar()
MCR_RL_RADAR_Status= IntVar()
MFR1_RADAR_Status= IntVar()

#Open app config canoe port
def Hardware_Config():
      subprocess.call('C:/Windows/System32/vcanconf.exe')  
#========================================Open database========================================		
def OpenDBC():    
    root.attributes('-topmost', True)
    open_file = filedialog.askopenfilename() # Returns opened path as str
    print(open_file) 
    db = cantools.database.load_file(open_file)
    for i in range(100):
        try:
            print(db.messages[i].name,db.messages[i].senders,db.messages[i].length,hex(db.messages[i].frame_id),db.messages[i].cycle_time,'\n')
        except IndexError:
            break
    showinfo(title='Selected File', message=open_file)

# Taoj excel file and save
def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)
#================================Select excel file==============================================
def select_filexlsx():
    filetypes = (
        ('text files', '*.xlsx'),
        ('All files', '*.*')
    )
    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)
    showinfo(
        title='Selected File',
        message=filename
    )
    return filename
#================================Select DBC file==============================================
def select_filedbc():
    filetypes = (
        ('text files', '*.dbc'),
        ('All files', '*.*')
    )
    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)
    showinfo(
        title='Selected File',
        message=filename
    )
    return filename
#======================================Read DID===========================================

#=============================Read DIDACM=================================================
def ACM_READ_DID():
    ECU_Text('ACM',None,1,None)
    ECU_Req_INFOR(ACM_DiagRq_HW,ACM_DiagRq,HW,0x608,'ACM','HardWare',0,'ACMF191')
    ECU_Req_INFOR(ACM_DiagRq_HW_Rv,ACM_DiagRq,HW_rv,0x608,'ACM','Hardware_Rev',0,'ACMF191')
    ECU_Req_INFOR(ACM_DiagRq_SW,ACM_DiagRq,SW,0x608,'ACM','Software',0,'ACMF188')
    ECU_Req_INFOR(ACM_DiagRq_SW_Rv,ACM_DiagRq,HW_rv,0x608,'ACM','Software_Rev',0,'ACMF188')
    ECU_Req_INFOR(ACM_DiagRq_CAL,ACM_DiagRq,SW,0x608,'ACM','Software_CAL',0,'ACMF102')
    ECU_Req_INFOR(ACM_DiagRq_CAL_Rev,ACM_DiagRq,HW_rv,0x608,'ACM','Software_CAL_Rev',0,'ACMF102')
    ECU_Req_INFOR(ACM_DiagRq_Bootloader,ACM_DiagRq,bl,0x608,'ACM','Bootloader',0,'ACMF101')
    bus.shutdown
#=============================Read DIDAPM================================
def APM_READ_DID():
    ECU_Text('APM',None,1,None)
    ECU_Req_INFOR(APM_DiagRq_HW,APM_DiagRq,HW,0x641,'APM','HardWare',0,'APMF191')
    ECU_Req_INFOR(APM_DiagRq_HW_Rv,APM_DiagRq,HW_rv,0x641,'APM','Hardware_Rev',0,'APMF191')
    ECU_Req_INFOR(APM_DiagRq_SW,APM_DiagRq,SW,0x641,'APM','Software',0,'APMF188')
    ECU_Req_INFOR(APM_DiagRq_SW_Rv,APM_DiagRq,HW_rv,0x641,'APM','Software_Rev',0,'APMF188')
    ECU_Req_INFOR(APM_DiagRq_Bootloader,APM_DiagRq,bl,0x641,'APM','Bootloader',0,'APMF101')
    bus.shutdown
#=============================Read DIDAVAS================================
def AVAS_READ_DID():
    ECU_Text('AVAS',None,1,None)
    ECU_Req_INFOR(AVAS_DiagRq_HW,AVAS_DiagRq,HW,0x62A,'AVAS','HardWare',0,'AVASF191')
    ECU_Req_INFOR(AVAS_DiagRq_HW_Rv,AVAS_DiagRq,HW_rv,0x62A,'AVAS','Hardware_Rev',0,'AVASF191')
    ECU_Req_INFOR(AVAS_DiagRq_SW,AVAS_DiagRq,SW,0x62A,'AVAS','Software',0,'AVASF188')
    ECU_Req_INFOR(AVAS_DiagRq_SW_Rv,AVAS_DiagRq,HW_rv,0x62A,'AVAS','Software_Rev',0,'AVASF188')
    ECU_Req_INFOR(AVAS_DiagRq_Bootloader,AVAS_DiagRq,bl,0x62A,'AVAS','Bootloader',0,'AVASF101')
    bus.shutdown
#=============================Read DIDBCM_BPM================================
def BCM_BPM_READ_DID():
    ECU_Text('BCM_BPM',None,1,None)
    ECU_Req_INFOR(BCM_BPM_DiagRq_HW,BCM_BPM_DiagRq,HW,0x67B,'BCM_BPM','HardWare',0,'BCM_BPMF191')
    ECU_Req_INFOR(BCM_BPM_DiagRq_HW_Rv,BCM_BPM_DiagRq,HW_rv,0x67B,'BCM_BPM','Hardware_Rev',0,'BCM_BPMF191')
    ECU_Req_INFOR(BCM_BPM_DiagRq_SW,BCM_BPM_DiagRq,SW,0x67B,'BCM_BPM','Software',0,'BCM_BPMF188')
    ECU_Req_INFOR(BCM_BPM_DiagRq_SW_Rv,BCM_BPM_DiagRq,HW_rv,0x67B,'BCM_BPM','Software_Rev',0,'BCM_BPMF188')
    bus.shutdown
#=============================Read DIDBCM================================
def BCM_READ_DID():    
    ECU_Text('BCM',None,1,None)
    ECU_Req_INFOR(BCM_DiagRq_HW,BCM_DiagRq,HW,0x601,'BCM','Hardware',0,'BCMF191')
    ECU_Req_INFOR(BCM_DiagRq_HW_Rv,BCM_DiagRq,HW_rv,0x601,'BCM','Hardware_Rev',0,'BCMF191')
    ECU_Req_INFOR(BCM_DiagRq_SW,BCM_DiagRq,SW,0x601,'BCM','Software',0,'BCMF188')
    ECU_Req_INFOR(BCM_DiagRq_SW_Rv,BCM_DiagRq,HW_rv,0x601,'BCM','Software_Rev',0,'BCMF188')
    ECU_Req_INFOR(BCM_DiagRq_Bootloader,BCM_DiagRq,bl,0x601,'BCM','Bootloader',0,'BCMF101')
    bus.shutdown
#=============================Read DIDBMS================================
def BMS_READ_DID():
    ECU_Text('BMS',None,1,None)
    ECU_Req_INFOR(BMS_DiagRq_HW,BMS_DiagRq,HW,0x613,'BMS','Hardware',0,'BMSF191')
    ECU_Req_INFOR(BMS_DiagRq_HW_Rv,BMS_DiagRq,HW_rv,0x613,'BMS','Hardware_Rev',0,'BMSF191')
    ECU_Req_INFOR(BMS_DiagRq_SW,BMS_DiagRq,SW,0x613,'BMS','Software',0,'BMSF188')
    ECU_Req_INFOR(BMS_DiagRq_SW_Rv,BMS_DiagRq,HW_rv,0x613,'BMS','Software_Rev',0,'BMSF188') 
    ECU_Req_INFOR(BMS_DiagRq_CAL,BMS_DiagRq,SW,0x613,'BMS','Software_CAL',0,'BMSF102')   
    ECU_Req_INFOR(BMS_DiagRq_CAL_Rev,BMS_DiagRq,HW_rv,0x613,'BMS','Software_CAL_Rev',0,'BMSF102')
    ECU_Req_INFOR(BMS_DiagRq_Bootloader,BMS_DiagRq,bl,0x613,'BMS','Bootloader',0,'BMSF101')
    bus.shutdown
#=============================Read DIDCCUF================================
def CCUF_READ_DID():
    ECU_Text('CCUF',None,1,None)
    ECU_Req_INFOR(CCUF_DiagRq_HW,CCUF_DiagRq,HW,0x609,'CCU_F','HardWare',0,'CCUFF191')
    ECU_Req_INFOR(CCUF_DiagRq_HW_Rv,CCUF_DiagRq,HW_rv,0x609,'CCU_F','Hardware_Rev',0,'CCUFF191')
    ECU_Req_INFOR(CCUF_DiagRq_SW,CCUF_DiagRq,SW,0x609,'CCU_F','Software',0,'CCUFF188')
    ECU_Req_INFOR(CCUF_DiagRq_SW_Rv,CCUF_DiagRq,HW_rv,0x609,'CCU_F','Software_Rev',0,'CCUFF188')  
    ECU_Req_INFOR(CCUF_DiagRq_CAL,CCUF_DiagRq,SW,0x609,'CCU_F','Software_CAL',0,'CCUFF102')
    ECU_Req_INFOR(CCUF_DiagRq_CAL_Rev,CCUF_DiagRq,HW_rv,0x609,'CCU_F','Software_CAL_Rev',0,'CCUFF102')
    ECU_Req_INFOR(CCUF_DiagRq_Bootloader,CCUF_DiagRq,bl,0x609,'CCU_F','Bootloader',0,'CCUFF101')
    bus.shutdown
#=============================Read DIDCPD================================
def CPD_READ_DID():
    ECU_Text('CPD',None,1,None)
    ECU_Req_INFOR(CPD_DiagRq_HW,CPD_DiagRq,HW,0x62F,'CPD','HardWare',0,'CPDF191')
    ECU_Req_INFOR(CPD_DiagRq_HW_Rv,CPD_DiagRq,HW_rv,0x62F,'CPD','Hardware_Rev',0,'CPDF191')
    ECU_Req_INFOR(CPD_DiagRq_SW,CPD_DiagRq,SW,0x62F,'CPD','Software',0,'CPDF191')
    ECU_Req_INFOR(CPD_DiagRq_SW_Rv,CPD_DiagRq,HW_rv,0x62F,'CPD','Software_Rev',0,'CPDF191') 
    ECU_Req_INFOR(CPD_DiagRq_Bootloader,CPD_DiagRq,bl,0x62F,'CPD','Bootloader',0,'CPDF191')
    bus.shutdown
#=============================Read DIDEDS_F================================
def EDS_F_READ_DID():
    ECU_Text('EDS_F',None,1,None)
    ECU_Req_INFOR(EDS_F_DiagRq_HW,EDS_F_DiagRq,HW,0x61B,'EDS_F','HardWare',0,'EDS_FF191')
    ECU_Req_INFOR(EDS_F_DiagRq_HW_Rv,EDS_F_DiagRq,HW_rv,0x61B,'EDS_F','Hardware_Rev',0,'EDS_FF191')
    ECU_Req_INFOR(EDS_F_DiagRq_SW,EDS_F_DiagRq,SW,0x61B,'EDS_F','Software',0,'EDS_FF188')
    ECU_Req_INFOR(EDS_F_DiagRq_SW_Rv,EDS_F_DiagRq,HW_rv,0x61B,'EDS_F','Software_Rev',0,'EDS_FF188')  
    ECU_Req_INFOR(EDS_F_DiagRq_CAL,EDS_F_DiagRq,SW,0x61B,'EDS_F','Software_CAL',0,'EDS_FF102')
    ECU_Req_INFOR(EDS_F_DiagRq_CAL_Rev,EDS_F_DiagRq,HW_rv,0x61B,'EDS_F','Software_CAL_Rev',0,'EDS_FF102')
    ECU_Req_INFOR(EDS_F_DiagRq_Bootloader,EDS_F_DiagRq,bl,0x61B,'EDS_F','Bootloader',0,'EDS_FF101')
    bus.shutdown
#=============================Read DIDEDS_R================================
def EDS_R_READ_DID():
    ECU_Text('EDS_R',None,1,None)
    ECU_Req_INFOR(EDS_R_DiagRq_HW,EDS_R_DiagRq,HW,0x61A,'EDS_R','HardWare',0,'EDS_RF191')
    ECU_Req_INFOR(EDS_R_DiagRq_HW_Rv,EDS_R_DiagRq,HW_rv,0x61A,'EDS_R','Hardware_Rev',0,'EDS_RF191')
    ECU_Req_INFOR(EDS_R_DiagRq_SW,EDS_R_DiagRq,SW,0x61A,'EDS_R','Software',0,'EDS_RF188')
    ECU_Req_INFOR(EDS_R_DiagRq_SW_Rv,EDS_R_DiagRq,HW_rv,0x61A,'EDS_R','Software_Rev',0,'EDS_RF188')
    ECU_Req_INFOR(EDS_R_DiagRq_CAL,EDS_R_DiagRq,SW,0x61A,'EDS_R','Software_CAL',0,'EDS_RF102')
    ECU_Req_INFOR(EDS_R_DiagRq_CAL_Rev,EDS_R_DiagRq,HW_rv,0x61A,'EDS_R','Software_CAL_Rev',0,'EDS_RF102')    
    ECU_Req_INFOR(EDS_R_DiagRq_Bootloader,EDS_R_DiagRq,bl,0x61A,'EDS_R','Bootloader',0,'EDS_RF101')
    bus.shutdown
#=============================Read DIDEPS================================
def EPS_READ_DID():
    ECU_Text('EPS',None,1,None)
    ECU_Req_INFOR(EPS_DiagRq_HW,EPS_DiagRq,HW,0x628,'EPS1','HardWare',0,'EPS1F191')
    ECU_Req_INFOR(EPS_DiagRq_HW_Rv,EPS_DiagRq,HW_rv,0x628,'EPS1','Hardware_Rev',0,'EPS1F191')
    ECU_Req_INFOR(EPS_DiagRq_SW,EPS_DiagRq,SW,0x628,'EPS1','Software',0,'EPS1F188') 
    ECU_Req_INFOR(EPS_DiagRq_SW_Rv,EPS_DiagRq,HW_rv,0x628,'EPS1','Software_Rev',0,'EPS1F188')
    ECU_Req_INFOR(EPS_DiagRq_Bootloader,EPS_DiagRq,bl,0x628,'EPS1','Bootloader',0,'EPS1F101')
    bus.shutdown
#=============================Read DIDETG================================
def ETG_READ_DID():
    ECU_Text('ETG',None,1,None)
    ECU_Req_INFOR(ETG_DiagRq_HW,ETG_DiagRq,HW,0x60E,'ETG','HardWare',0,'ETGF191')
    ECU_Req_INFOR(ETG_DiagRq_HW_Rv,ETG_DiagRq,HW_rv,0x60E,'ETG','Hardware_Rev',0,'ETGF191')
    ECU_Req_INFOR(ETG_DiagRq_SW,ETG_DiagRq,SW,0x60E,'ETG','Software',0,'ETGF188')
    ECU_Req_INFOR(ETG_DiagRq_SW_Rv,ETG_DiagRq,HW_rv,0x60E,'ETG','Software_Rev',0,'ETGF188')
    ECU_Req_INFOR(ETG_DiagRq_Bootloader,ETG_DiagRq,bl,0x60E,'ETG','Bootloader',0,'ETGF101')
    bus.shutdown
#=============================Read DIDFCAM================================
def FCAM_READ_DID():
    ECU_Text('FCAM',None,1,None)
    ECU_Req_INFOR(FCAM_DiagRq_HW,FCAM_DiagRq,HW,0x616,'SCAM','HardWare',0,'SCAMF191')
    ECU_Req_INFOR(FCAM_DiagRq_HW_Rv,FCAM_DiagRq,HW_rv,0x616,'SCAM','Hardware_Rev',0,'SCAMF191')
    ECU_Req_INFOR(FCAM_DiagRq_SW,FCAM_DiagRq,SW,0x616,'SCAM','Software',0,'SCAMF188')
    ECU_Req_INFOR(FCAM_DiagRq_SW_Rv,FCAM_DiagRq,HW_rv,0x616,'SCAM','Software_Rev',0,'SCAMF188')
    ECU_Req_INFOR(FCAM_DiagRq_CAL,FCAM_DiagRq,SW,0x616,'SCAM','Software_CAL',0,'SCAMF102')
    ECU_Req_INFOR(FCAM_DiagRq_CAL_Rev,FCAM_DiagRq,HW_rv,0x616,'SCAM','Software_CAL_Rev',0,'SCAMF102')
    ECU_Req_INFOR(FCAM_DiagRq_Bootloader,FCAM_DiagRq,bl,0x616,'SCAM','Bootloader',0,'SCAMF101')
    bus.shutdown
#=============================Read DIDGS================================
def GS_READ_DID():
    ECU_Text('GS',None,1,None)
    ECU_Req_INFOR(GS_DiagRq_HW,GS_DiagRq,HW,0x610,'GS','HardWare',0,'GSF191')
    ECU_Req_INFOR(GS_DiagRq_HW_Rv,GS_DiagRq,HW_rv,0x610,'GS','Hardware_Rev',0,'GSF191')
    ECU_Req_INFOR(GS_DiagRq_SW,GS_DiagRq,SW,0x610,'GS','Software',0,'GSF188') 
    ECU_Req_INFOR(GS_DiagRq_SW_Rv,GS_DiagRq,HW_rv,0x610,'GS','Software_Rev',0,'GSF188')
    ECU_Req_INFOR(GS_DiagRq_Bootloader,GS_DiagRq,bl,0x610,'GS','Bootloader',0,'GSF101')
    bus.shutdown
#=============================Read DIDHUD================================
def HUD_READ_DID():
    ECU_Text('HUD',None,1,None)
    ECU_Req_INFOR(HUD_DiagRq_HW,HUD_DiagRq,HW,0x60D,'HUD','HardWare',0,'HUDF191')
    ECU_Req_INFOR(HUD_DiagRq_HW_Rv,HUD_DiagRq,HW_rv,0x60D,'HUD','Hardware_Rev',0,'HUDF191')
    ECU_Req_INFOR(HUD_DiagRq_SW_APP,HUD_DiagRq,SW,0x60D,'HUD','Software_APP',0,'HUDF188')
    ECU_Req_INFOR(HUD_DiagRq_SW_APP_Rv,HUD_DiagRq,HW_rv,0x60D,'HUD','Software_APP_Rev',0,'HUDF188')  
    ECU_Req_INFOR(HUD_DiagRq_SW_BASIC,HUD_DiagRq,SW,0x60D,'HUD','Software_BASIC',0,'HUDF105')
    ECU_Req_INFOR(HUD_DiagRq_SW_BASIC_Rv,HUD_DiagRq,HW_rv,0x60D,'HUD','Software_BASIC_Rev',0,'HUDF105') 
    ECU_Req_INFOR(HUD_DiagRq_HMI,HUD_DiagRq,SW,0x60D,'HUD','Software_HMI',0,'HUDF106')
    ECU_Req_INFOR(HUD_DiagRq_HMI_Rv,HUD_DiagRq,HW_rv,0x60D,'HUD','Software_HMI_Rev',0,'HUDF106') 
    ECU_Req_INFOR(HUD_DiagRq_Bootloader,HUD_DiagRq,bl,0x60D,'HUD','Bootloader',0,'HUDF101')
    bus.shutdown
#=============================Read DIDIDB================================
def IDB_READ_DID():
    ECU_Text('IDB',None,1,None)
    ECU_Req_INFOR(IDB_DiagRq_HW,IDB_DiagRq,HW,0x629,'IDB','HardWare',0,'IDBF191')
    ECU_Req_INFOR(IDB_DiagRq_HW_Rv,IDB_DiagRq,HW_rv,0x629,'IDB','Hardware_Rev',0,'IDBF191')
    ECU_Req_INFOR(IDB_DiagRq_SW,IDB_DiagRq,SW,0x629,'IDB','Software',0,'IDBF188') 
    ECU_Req_INFOR(IDB_DiagRq_SW_Rv,IDB_DiagRq,HW_rv,0x629,'IDB','Software_Rev',0,'IDBF188')
    ECU_Req_INFOR(IDB_DiagRq_CAL,IDB_DiagRq,SW,0x629,'IDB','Software',0,'IDBF102')
    ECU_Req_INFOR(IDB_DiagRq_CAL_Rev,IDB_DiagRq,HW_rv,0x629,'IDB','Software_CAL_Rev',0,'IDBF102')
    ECU_Req_INFOR(IDB_DiagRq_Bootloader,IDB_DiagRq,bl,0x629,'IDB','Bootloader',0,'IDBF101')
    bus.shutdown
#=============================Read DIDMHU================================
def MHU_READ_DID():
    ECU_Text('MHU',None,1,None)    
    ECU_Req_INFOR(MHU_DiagRq_HW,MHU_DiagRq,HW,0x604,'MHU','HardWare',0,'MHUF191')
    ECU_Req_INFOR(MHU_DiagRq_HW_Rv,MHU_DiagRq,HW_rv,0x604,'MHU','Hardware_Rev',0,'MHUF191')
    ECU_Req_INFOR(MHU_DiagRq_SW,MHU_DiagRq,SW,0x604,'MHU','Software',0,'MHUF188')
    ECU_Req_INFOR(MHU_DiagRq_SW_Rv,MHU_DiagRq,HW_rv,0x604,'MHU','Software_Rev',0,'MHUF188')
    ECU_Req_INFOR(MHU_DiagRq_TBOX,MHU_DiagRq,SW,0x604,'MHU','Software_TBOX',0,'MHUF164')
    ECU_Req_INFOR(MHU_DiagRq_TBOX_Rv,MHU_DiagRq,HW_rv,0x604,'MHU','Software_TBOX_Rev',0,'MHUF164')
    ECU_Req_INFOR(MHU_DiagRq_Bootloader,MHU_DiagRq,bl,0x604,'MHU','Bootloader',0,'MHUF101')
    bus.shutdown
#=============================Read DIDMRGEN================================
def MRGEN_READ_DID():
    ECU_Text('MRGEN',None,1,None)  
    ECU_Req_INFOR(MRGEN_DiagRq_HW,MRGEN_DiagRq,HW,0x625,'MRGEN','HardWare',0,'MRGENF191')
    ECU_Req_INFOR(MRGEN_DiagRq_HW_Rv,MRGEN_DiagRq,HW_rv,0x625,'MRGEN','Hardware_Rev',0,'MRGENF191')
    ECU_Req_INFOR(MRGEN_DiagRq_SW,MRGEN_DiagRq,SW,0x625,'MRGEN','Software',0,'MRGENF188')
    ECU_Req_INFOR(MRGEN_DiagRq_SW_Rv,MRGEN_DiagRq,HW_rv,0x625,'MRGEN','Software_Rev',0,'MRGENF188')
    ECU_Req_INFOR(MRGEN_DiagRq_Bootloader,MRGEN_DiagRq,bl,0x625,'MRGEN','Bootloader',0,'MRGENF101')
    bus.shutdown
#=============================Read DIDOCS_P================================
def OCS_P_READ_DID():
    ECU_Text('OCS',None,1,None)  
    ECU_Req_INFOR(OCS_P_DiagRq_HW,OCS_P_DiagRq,HW,0x644,'OCS','HardWare',0,'OCSF191')
    ECU_Req_INFOR(OCS_P_DiagRq_HW_Rv,OCS_P_DiagRq,HW_rv,0x644,'OCS','Hardware_Rev',0,'OCSF191')
    ECU_Req_INFOR(OCS_P_DiagRq_SW,OCS_P_DiagRq,SW,0x644,'OCS','Software',0,'OCSF188')
    ECU_Req_INFOR(OCS_P_DiagRq_SW_Rv,OCS_P_DiagRq,HW_rv,0x644,'OCS','Software_Rev',0,'OCSF188')
    ECU_Req_INFOR(OCS_P_DiagRq_Bootloader,OCS_P_DiagRq,bl,0x644,'OCS','Bootloader',0,'OCSF101')
    bus.shutdown
#=============================Read DIDPAS================================
def PAS_READ_DID():
    ECU_Text('AP',None,1,None)  
    ECU_Req_INFOR(PAS_DiagRq_HW,PAS_DiagRq,HW,0x617,'ADAS','HardWare',0,'ADASF191')
    ECU_Req_INFOR(PAS_DiagRq_HW_Rv,PAS_DiagRq,HW_rv,0x617,'ADAS','Hardware_Rev',0,'ADASF191')
    ECU_Req_INFOR(PAS_DiagRq_SW,PAS_DiagRq,SW,0x617,'ADAS','Software',0,'ADASF188')
    ECU_Req_INFOR(PAS_DiagRq_SW_Rv,PAS_DiagRq,HW_rv,0x617,'ADAS','Software_Rev',0,'ADASF188')
    ECU_Req_INFOR(PAS_DiagRq_Bootloader,PAS_DiagRq,bl,0x617,'ADAS','Bootloader',0,'ADASF101')
    bus.shutdown
#=============================Read DIDPOD_DCDC================================
def POD_DCDC_READ_DID():
    ECU_Text('POD_DCDC',None,1,None)  
    ECU_Req_INFOR(POD_DCDC_DiagRq_HW,POD_DCDC_DiagRq,HW,0x63D,'DCDC','HardWare',0,'DCDCF191')
    ECU_Req_INFOR(POD_DCDC_DiagRq_HW_Rv,POD_DCDC_DiagRq,HW_rv,0x63D,'DCDC','Hardware_Rev',0,'DCDCF191')
    ECU_Req_INFOR(POD_DCDC_DiagRq_SW,POD_DCDC_DiagRq,SW,0x63D,'DCDC','Software',0,'DCDCF188')
    ECU_Req_INFOR(POD_DCDC_DiagRq_SW_Rv,POD_DCDC_DiagRq,HW_rv,0x63D,'DCDC','Software_Rev',0,'DCDCF188')
    ECU_Req_INFOR(POD_DCDC_DiagRq_CAL,POD_DCDC_DiagRq,SW,0x63D,'DCDC','Software_CAL',0,'DCDCF102')
    ECU_Req_INFOR(POD_DCDC_DiagRq_CAL_Rev,POD_DCDC_DiagRq,HW_rv,0x63D,'DCDC','Software_CAL_Rev',0,'DCDCF102')
    ECU_Req_INFOR(POD_DCDC_DiagRq_Bootloader,POD_DCDC_DiagRq,bl,0x63D,'DCDC','Bootloader',0,'DCDCF101')
    bus.shutdown
#=============================Read DIDPOD_GW================================
def POD_GW_READ_DID():
    ECU_Text('POD_GW',None,1,None)  
    ECU_Req_INFOR(POD_GW_DiagRq_HW,POD_GW_DiagRq,HW,0x63F,'POD_GW','HardWare',0,'POD_GWF191')
    ECU_Req_INFOR(POD_GW_DiagRq_HW_Rv,POD_GW_DiagRq,HW_rv,0x63F,'POD_GW','Hardware_Rev',0,'POD_GWF191')
    ECU_Req_INFOR(POD_GW_DiagRq_SW,POD_GW_DiagRq,SW,0x63F,'POD_GW','Software',0,'POD_GWF188')
    ECU_Req_INFOR(POD_GW_DiagRq_SW_Rv,POD_GW_DiagRq,HW_rv,0x63F,'POD_GW','Software_Rev',0,'POD_GWF188')
    ECU_Req_INFOR(POD_GW_DiagRq_Bootloader,POD_GW_DiagRq,bl,0x63F,'POD_GW','Bootloader',0,'POD_GWF101')
    bus.shutdown
#=============================Read DIDPOD_OBC================================
def POD_OBC_READ_DID():
    ECU_Text('POD_OBC',None,1,None)
    ECU_Req_INFOR(POD_OBC_DiagRq_HW,POD_OBC_DiagRq,HW,0x63E,'OBC','HardWare',0,'OBCF191')
    ECU_Req_INFOR(POD_OBC_DiagRq_HW_Rv,POD_OBC_DiagRq,HW_rv,0x63E,'OBC','Hardware_Rev',0,'OBCF191')
    ECU_Req_INFOR(POD_OBC_DiagRq_SW,POD_OBC_DiagRq,SW,0x63E,'OBC','Software',0,'OBCF188')
    ECU_Req_INFOR(POD_OBC_DiagRq_SW_Rv,POD_OBC_DiagRq,HW_rv,0x63E,'OBC','Software_Rev',0,'OBCF188')
    ECU_Req_INFOR(POD_OBC_DiagRq_CAL,POD_OBC_DiagRq,SW,0x63E,'OBC','Software_CAL',0,'OBCF102')
    ECU_Req_INFOR(POD_OBC_DiagRq_CAL_Rev,POD_OBC_DiagRq,HW_rv,0x63E,'OBC','Software_CAL_Rev',0,'OBCF102')
    ECU_Req_INFOR(POD_OBC_DiagRq_Bootloader,POD_OBC_DiagRq,bl,0x63E,'OBC','Bootloader',0,'OBCF101')
    bus.shutdown
#=============================Read DIDRCU================================
def RCU_READ_DID():
    ECU_Text('RCU',None,1,None)
    ECU_Req_INFOR(RCU_DiagRq_HW,RCU_DiagRq,HW,0x627,'RCU','HardWare',0,'RCUF191')
    ECU_Req_INFOR(RCU_DiagRq_HW_Rv,RCU_DiagRq,HW_rv,0x627,'RCU','Hardware_Rev',0,'RCUF191')
    ECU_Req_INFOR(RCU_DiagRq_SW,RCU_DiagRq,SW,0x627,'RCU','Software',0,'RCUF191')
    ECU_Req_INFOR(RCU_DiagRq_SW_Rv,RCU_DiagRq,HW_rv,0x627,'RCU','Software_Rev',0,'RCUF191')
    ECU_Req_INFOR(RCU_DiagRq_CAL,RCU_DiagRq,SW,0x627,'RCU','Software_CAL',0,'RCUF102')
    ECU_Req_INFOR(RCU_DiagRq_CAL_Rev,RCU_DiagRq,HW_rv,0x627,'RCU','Software_CAL_Rev',0,'RCUF102')
    ECU_Req_INFOR(RCU_DiagRq_Bootloader,RCU_DiagRq,bl,0x627,'RCU','Bootloader',0,'RCUF101')
    bus.shutdown
#=============================Read DIDSHVU_F================================
def SHVU_F_READ_DID():
    ECU_Text('SHVU_F',None,1,None)
    ECU_Req_INFOR(SHVU_F_DiagRq_HW,SHVU_F_DiagRq,HW,0x642,'SHVU_F','HardWare',0,'SHVU_FF191')
    ECU_Req_INFOR(SHVU_F_DiagRq_HW_Rv,SHVU_F_DiagRq,HW_rv,0x642,'SHVU_F','Hardware_Rev',0,'SHVU_FF191')
    ECU_Req_INFOR(SHVU_F_DiagRq_SW,SHVU_F_DiagRq,SW,0x642,'SHVU_F','Software',0,'SHVU_FF188')
    ECU_Req_INFOR(SHVU_F_DiagRq_SW_Rv,SHVU_F_DiagRq,HW_rv,0x642,'SHVU_F','Software_Rev',0,'SHVU_FF188')
    ECU_Req_INFOR(SHVU_F_DiagRq_Bootloader,SHVU_F_DiagRq,bl,0x642,'SHVU_F','Bootloader',0,'SHVU_FF101')
    bus.shutdown
#=============================Read DIDSHVU_R================================
def SHVU_R_READ_DID():
    ECU_Text('SHVU_R',None,1,None)
    ECU_Req_INFOR(SHVU_R_DiagRq_HW,SHVU_R_DiagRq,HW,0x643,'SHVU_R','HardWare',0,'SHVU_RF191')
    ECU_Req_INFOR(SHVU_R_DiagRq_HW_Rv,SHVU_R_DiagRq,HW_rv,0x643,'SHVU_R','Hardware_Rev',0,'SHVU_RF191')
    ECU_Req_INFOR(SHVU_R_DiagRq_SW,SHVU_R_DiagRq,SW,0x643,'SHVU_R','Software',0,'SHVU_RF188')
    ECU_Req_INFOR(SHVU_R_DiagRq_SW_Rv,SHVU_R_DiagRq,HW_rv,0x643,'SHVU_R','Software_Rev',0,'SHVU_RF188')
    ECU_Req_INFOR(SHVU_R_DiagRq_Bootloader,SHVU_R_DiagRq,bl,0x643,'SHVU_R','Bootloader',0,'SHVU_RF101')
    bus.shutdown
#=============================Read DIDSRR_FL================================
def SRR_FL_READ_DID():
    ECU_Text('SRR_FL',None,1,None)
    ECU_Req_INFOR(SRR_FL_DiagRq_HW,SRR_FL_DiagRq,HW,0x635,'SRR_FL','HardWare',0,'SRR_FLF191')
    ECU_Req_INFOR(SRR_FL_DiagRq_HW_Rv,SRR_FL_DiagRq,HW_rv,0x635,'SRR_FL','Hardware_Rev',0,'SRR_FLF191')
    ECU_Req_INFOR(SRR_FL_DiagRq_SW,SRR_FL_DiagRq,SW,0x635,'SRR_FL','Software',0,'SRR_FLF199')
    ECU_Req_INFOR(SRR_FL_DiagRq_SW_Rv,SRR_FL_DiagRq,HW_rv,0x635,'SRR_FL','Software_Rev',0,'SRR_FLF199')
    ECU_Req_INFOR(SRR_FL_DiagRq_Bootloader,SRR_FL_DiagRq,bl,0x635,'SRR_FL','Bootloader',0,'SRR_FLF101')
    bus.shutdown
#=============================Read DIDSRR_FR================================
def SRR_FR_READ_DID():
    ECU_Text('SRR_FR',None,1,None)
    ECU_Req_INFOR(SRR_FR_DiagRq_HW,SRR_FR_DiagRq,HW,0x634,'SRR_FR','HardWare',0,'SRR_FRF191')
    ECU_Req_INFOR(SRR_FR_DiagRq_HW_Rv,SRR_FR_DiagRq,HW_rv,0x634,'SRR_FR','Hardware_Rev',0,'SRR_FRF191')
    ECU_Req_INFOR(SRR_FR_DiagRq_SW,SRR_FR_DiagRq,SW,0x634,'SRR_FR','Software',0,'SRR_FRF188')
    ECU_Req_INFOR(SRR_FR_DiagRq_SW_Rv,SRR_FR_DiagRq,HW_rv,0x634,'SRR_FR','Software_Rev',0,'SRR_FRF188')
    ECU_Req_INFOR(SRR_FR_DiagRq_Bootloader,SRR_FR_DiagRq,bl,0x634,'SRR_FR','Bootloader',0,'SRR_FRF101')
    bus.shutdown
#=============================Read DIDSRR_RL================================
def SRR_RL_READ_DID():
    ECU_Text('SRR_RL',None,1,None)
    ECU_Req_INFOR(SRR_RL_DiagRq_HW,SRR_RL_DiagRq,HW,0x637,'SRR_RL','HardWare',0,'SRR_RLF191')
    ECU_Req_INFOR(SRR_RL_DiagRq_HW_Rv,SRR_RL_DiagRq,HW_rv,0x637,'SRR_RL','Hardware_Rev',0,'SRR_RLF191')
    ECU_Req_INFOR(SRR_RL_DiagRq_SW,SRR_RL_DiagRq,SW,0x637,'SRR_RL','Software',0,'SRR_RLF181')
    ECU_Req_INFOR(SRR_RL_DiagRq_SW_Rv,SRR_RL_DiagRq,HW_rv,0x637,'SRR_RL','Software_Rev',0,'SRR_RLF181')
    ECU_Req_INFOR(SRR_RL_DiagRq_Bootloader,SRR_RL_DiagRq,bl,0x637,'SRR_RL','Bootloader',0,'SRR_RLF101')
    bus.shutdown
#=============================Read DIDSRR_RR================================
def SRR_RR_READ_DID():
    ECU_Text('SRR_RR',None,1,None)
    ECU_Req_INFOR(SRR_RR_DiagRq_HW,SRR_RR_DiagRq,HW,0x636,'SRR_RR','HardWare',0,'SRR_RRF191')
    ECU_Req_INFOR(SRR_RR_DiagRq_HW_Rv,SRR_RR_DiagRq,HW_rv,0x636,'SRR_RR','Hardware_Rev',0,'SRR_RRF191')
    ECU_Req_INFOR(SRR_RR_DiagRq_SW,SRR_RR_DiagRq,SW,0x636,'SRR_RR','Software',0,'SRR_RRF188')
    ECU_Req_INFOR(SRR_RR_DiagRq_SW_Rv,SRR_RR_DiagRq,HW_rv,0x636,'SRR_RR','Software_Rev',0,'SRR_RRF188')
    ECU_Req_INFOR(SRR_RR_DiagRq_Bootloader,SRR_RR_DiagRq,bl,0x636,'SRR_RR','Bootloader',0,'SRR_RRF101')
    bus.shutdown
#=============================Read DIDVCU================================
def VCU_READ_DID():
    ECU_Text('VCU',None,1,None)
    ECU_Req_INFOR(VCU_DiagRq_HW,VCU_DiagRq,HW,0x62C,'VCU','HardWare',0,'VCUF191')
    ECU_Req_INFOR(VCU_DiagRq_HW_Rv,VCU_DiagRq,HW_rv,0x62C,'VCU','Hardware_Rev',0,'VCUF191')
    ECU_Req_INFOR(VCU_DiagRq_SW,VCU_DiagRq,SW,0x62C,'VCU','Software',0,'VCUF188')
    ECU_Req_INFOR(VCU_DiagRq_SW_Rv,VCU_DiagRq,HW_rv,0x62C,'VCU','Software_Rev',0,'VCUF188')
    ECU_Req_INFOR(VCU_DiagRq_SW_APP,VCU_DiagRq,SW,0x62C,'VCU','Software_APP',0,'VCUF104')
    ECU_Req_INFOR(VCU_DiagRq_SW_APP_Rv,VCU_DiagRq,HW_rv,0x62C,'VCU','Software_APP_Rev',0,'VCUF104')
    ECU_Req_INFOR(VCU_DiagRq_SW_BASIC,VCU_DiagRq,SW,0x62C,'VCU','Software_BASIC',0,'VCUF105')
    ECU_Req_INFOR(VCU_DiagRq_SW_BASIC_Rv,VCU_DiagRq,HW_rv,0x62C,'VCU','Software_BASIC_Rev',0,'VCUF105')
    ECU_Req_INFOR(VCU_DiagRq_CAL,VCU_DiagRq,SW,0x62C,'VCU','Software_CAL',0,'VCUF102')
    ECU_Req_INFOR(VCU_DiagRq_CAL_Rv,VCU_DiagRq,HW_rv,0x62C,'VCU','Software_CAL_Rv',0,'VCUF102')
    ECU_Req_INFOR(VCU_DiagRq_Bootloader,VCU_DiagRq,bl,0x62C,'VCU','Bootloader',0,'VCUF101')
    bus.shutdown
#=============================Read DIDXGW================================
def XGW_READ_DID():
    ECU_Text('XGW',None,1,None)
    ECU_Req_INFOR(XGW_DiagRq_HW,XGW_DiagRq,HW,0x602,'XGW','Hardware',0,'XGWF191')
    ECU_Req_INFOR(XGW_DiagRq_HW_Rv,XGW_DiagRq,HW_rv,0x602,'XGW','Hardware_Rev',0,'XGWF191')
    ECU_Req_INFOR(XGW_DiagRq_SW,XGW_DiagRq,SW,0x602,'XGW','Software',0,'XGWF188')
    ECU_Req_INFOR(XGW_DiagRq_SW_Rv,XGW_DiagRq,HW_rv,0x602,'XGW','Software_Rev',0,'XGWF188')
    ECU_Req_INFOR(XGW_DiagRq_CAL,XGW_DiagRq,SW,0x602,'XGW','CAL',0,'XGWF102')
    ECU_Req_INFOR(XGW_DiagRq_CAL_Rv,XGW_DiagRq,HW_rv,0x602,'XGW','CAL_Rev',0,'XGWF102')
    ECU_Req_INFOR(XGW_DiagRq_Bootloader,XGW_DiagRq,bl,0x602,'XGW','Bootloader',0,'XGWF101')
    bus.shutdown
#=============================Read DIDYSS================================
#========================Read HW XGW=======================
def ECU_Req_INFOR(message1,message2,readstate,ECU_ID,ECU,TypeRead,Header_Text,iid):
    try:
        bus.send(message1)
        ECU_Resp(message2,readstate,ECU_ID,ECU,TypeRead,Header_Text,iid)
    except can.CanError:
        print("Msg cannot send")
#===========================================================
def ECU_Req(message,readstate,ECU_ID,ECU,TypeRead,Header_Text,iid):
    try:
        bus.send(message)
        ECU_Resp_Data(readstate,ECU_ID,ECU,TypeRead,Header_Text,iid)
    except can.CanError:
        print("Msg cannot send")
#===========================================================
def ECU_Resp_Data(readstate,ECU_ID,ECU,TypeRead,Header_Text,iid):
    DID_Infor = 'NRC'
    mess = bus.recv(timeout = 2)
    try:
        if mess.arbitration_id == ECU_ID:  
            if mess.data[0] == 33:
                if readstate == 1:
                    print("EEP",hex(mess.data[1])[2:],hex(mess.data[2])[2:],hex(mess.data[3])[2:],hex(mess.data[4])[2:])
                    DID_Infor = "EEP" + str(hex(mess.data[1])[2:2]) + str(hex(mess.data[2])[2:2]) + str(hex(mess.data[3])[2:2]) + str(hex(mess.data[4])[2:2])
                    ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
                    RealData(iid,DID_Infor,readstate)
                if readstate == 2:
                    print("SOW",hex(mess.data[1])[2:],hex(mess.data[2])[2:],hex(mess.data[3])[2:],hex(mess.data[4])[2:])
                    DID_Infor = "SOW" + str(hex(mess.data[1])[2:]) + str(hex(mess.data[2])[2:]) + str(hex(mess.data[3])[2:]) + str(hex(mess.data[4])[2:])
                    ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
                    RealData(iid,DID_Infor,readstate)
        else :
            DID_Infor = "NRC"
            ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
            RealData(iid,DID_Infor,readstate)
    except AttributeError:
        DID_Infor = "NRC"
        ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
        RealData(iid,DID_Infor,readstate)
    time.sleep(0.5)
#===========================================================
def ECU_Resp(message,readstate,ECU_ID,ECU,TypeRead,Header_Text,iid):
    DID_Infor = 'NRC'
    mess = bus.recv(timeout = 2)
    try:
        if mess.arbitration_id == ECU_ID:  
                    if mess.data[0] == 16 or mess.data[0] == 3:
                        ECU_Req(message,readstate,ECU_ID,ECU,TypeRead,Header_Text,iid)                    
                    if  mess.data[0] == 4:
                        print("REV",hex(mess.data[4])[2:])
                        DID_Infor = str(hex(mess.data[4])[2:])
                        ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
                        RealData(iid,DID_Infor,readstate)
                    if  mess.data[0] == 5:
                        print("Bld",hex(mess.data[4])[2:],hex(mess.data[5])[2:])
                        DID_Infor = str(hex(mess.data[4])[2:]) + str(hex(mess.data[5])[2:])
                        ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
                        RealData(iid,DID_Infor,readstate)
        else:
            DID_Infor = "NRC"
            ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
            RealData(iid,DID_Infor,readstate)
    except AttributeError:
        DID_Infor = "NRC"
        ECU_Text(ECU,TypeRead,Header_Text,DID_Infor)
        RealData(iid,DID_Infor,readstate)
    time.sleep(0.5)
#====================================VF3 Window============================================		
def VF3Window():     
    VF3 = Toplevel(root)
    VF3.title("VF3 DID")
    VF3.geometry("1200x500") 
#====================================VF5 Window============================================		
def VF5Window():
    newWindow = Toplevel(root)
    newWindow.title("New Window")
    newWindow.geometry("200x200") 
    T = Text(newWindow, height = 2, width = 5)
    T.grid(pady=50,padx=50)
    create_workbook("hello.xlsx")
#====================================VF6 Window============================================		
def VF6Window():
    VF6 = Toplevel(root)
    VF6.title("New Window")
    VF6.geometry("1200x600") 
    VF6.resizable(0, 0)
    Label(VF6, text ="VF6 DID").pack()
    global ECU_Text
    global RealData

    ECU_Frame = tk.LabelFrame(VF6,bg = '#c3c3c3',text= 'ECU',borderwidth=1,relief='solid')
    ECU_Frame.pack(side=tk.LEFT)
    ECU_Frame.pack_propagate(False)
    ECU_Frame.configure(width=100,height=900)

    Excel_frame = tk.LabelFrame(VF6,bg = '#c3c3c3',text= 'ECU Infor',borderwidth=1,relief='solid')
    Excel_frame.pack(side=tk.TOP,padx=1,pady=1)
    Excel_frame.pack_propagate(False)
    Excel_frame.configure(width=1400,height=400)

    style = ttk.Style()
    style.theme_use('clam')
    tree = ttk.Treeview(Excel_frame, column=("c1", "c2","c3","c4","c5","c6","c7","c8","c9","C10"), show='headings', height=8)
    tree.pack(expand=True, fill='y')

    Terminal_frame = tk.LabelFrame(VF6,text= 'Terminal',bg = '#c3c3c3',borderwidth=1,relief='solid')
    Terminal_frame.pack(side=tk.BOTTOM,padx=1,pady=1)
    Terminal_frame.pack_propagate(False)
    Terminal_frame.configure(width=1400,height=740)

    scrollbar = ScrolledText(ECU_Frame)
    scrollbar.pack( side = RIGHT, fill=Y )

    text_box = Text(Terminal_frame,height=12,width=150)
    text_box.pack(expand=True)

    sb = Scrollbar(Terminal_frame)
    sb.pack(side=RIGHT, fill=BOTH)

    text_box.config(yscrollcommand=sb.set)
    sb.config(command=text_box.yview)

    Terminal_frame.pack(expand=True)  
    def Click_Select():
        if ALL_Status.get() == 1:
            VCU_Status.set(1)
            DCDC_Status.set(1)
            POD_Status.set(1)
            OBC_Status.set(1)
            EDS_Status.set(1)
            BMS_Status.set(1)
            GS_Status.set(1)
            IDB_Status.set(1)
            RCU_Status.set(1)
            EPS_Status.set(1)
            ACM_Status.set(1)
            BCM_Status.set(1)
            BCM_BPM_Status.set(1)
            CCU1_Status.set(1)
            XGW_Status.set(1)
            APM_Status.set(1)
            SHVU_F_Status.set(1)
            SHVU_R_Status.set(1)
            OCS_Status.set(1)
            MHU_Status.set(1)
            HUD_Status.set(1)
            AVAS_Status.set(1)
            AP_ECU_Status.set(1)
            FCAM_Status.set(1)
            MCR_FL_RADAR_Status.set(1)
            MCR_FR_RADAR_Status.set(1)
            MCR_RR_RADAR_Status.set(1)
            MCR_RL_RADAR_Status.set(1)
            MFR1_RADAR_Status.set(1)
        else:
            VCU_Status.set(0)
            DCDC_Status.set(0)
            POD_Status.set(0)
            OBC_Status.set(0)
            EDS_Status.set(0)
            BMS_Status.set(0)
            GS_Status.set(0)
            IDB_Status.set(0)
            RCU_Status.set(0)
            EPS_Status.set(0)
            ACM_Status.set(0)
            BCM_Status.set(0)
            BCM_BPM_Status.set(0)
            CCU1_Status.set(0)
            XGW_Status.set(0)
            APM_Status.set(0)
            SHVU_F_Status.set(0)
            SHVU_R_Status.set(0)
            OCS_Status.set(0)
            MHU_Status.set(0)
            HUD_Status.set(0)
            AVAS_Status.set(0)
            AP_ECU_Status.set(0)
            FCAM_Status.set(0)
            MCR_FL_RADAR_Status.set(0)
            MCR_FR_RADAR_Status.set(0)
            MCR_RR_RADAR_Status.set(0)
            MCR_RL_RADAR_Status.set(0)
            MFR1_RADAR_Status.set(0)  

    ALL = tk.Checkbutton(scrollbar, text='ALL', bg='white', anchor='w',variable= ALL_Status,onvalue=1, offvalue=0,command= Click_Select)
    scrollbar.window_create('end', window=ALL)
    scrollbar.insert('end', '\n')
    VCU = tk.Checkbutton(scrollbar, text='VCU', bg='white', anchor='w',variable= VCU_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=VCU)
    scrollbar.insert('end', '\n')
    DCDC = tk.Checkbutton(scrollbar, text='DCDC', bg='white', anchor='w',variable= DCDC_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=DCDC)
    scrollbar.insert('end', '\n')
    POD = tk.Checkbutton(scrollbar, text='POD', bg='white', anchor='w',variable= POD_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=POD)
    scrollbar.insert('end', '\n')
    OBC = tk.Checkbutton(scrollbar, text='OBC', bg='white', anchor='w',variable= OBC_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=OBC)
    scrollbar.insert('end', '\n')
    EDS = tk.Checkbutton(scrollbar, text='EDS', bg='white', anchor='w',variable= EDS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=EDS)
    scrollbar.insert('end', '\n')
    BMS = tk.Checkbutton(scrollbar, text='BMS', bg='white', anchor='w',variable= BMS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=BMS)
    scrollbar.insert('end', '\n')
    GS = tk.Checkbutton(scrollbar, text='GS', bg='white', anchor='w',variable= GS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=GS)
    scrollbar.insert('end', '\n')
    IDB = tk.Checkbutton(scrollbar, text='IDB', bg='white', anchor='w',variable= IDB_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=IDB)
    scrollbar.insert('end', '\n')
    RCU = tk.Checkbutton(scrollbar, text='RCU', bg='white', anchor='w',variable= RCU_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=RCU)
    scrollbar.insert('end', '\n')
    EPS = tk.Checkbutton(scrollbar, text='EPS', bg='white', anchor='w',variable= EPS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=EPS)
    scrollbar.insert('end', '\n')
    ACM = tk.Checkbutton(scrollbar, text='ACM', bg='white', anchor='w',variable= ACM_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=ACM)
    scrollbar.insert('end', '\n')
    BCM = tk.Checkbutton(scrollbar, text='BCM', bg='white', anchor='w',variable= BCM_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=BCM)
    scrollbar.insert('end', '\n')
    BPM = tk.Checkbutton(scrollbar, text='BPM', bg='white', anchor='w',variable= BCM_BPM_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=BPM)
    scrollbar.insert('end', '\n')
    CCU = tk.Checkbutton(scrollbar, text='CCU', bg='white', anchor='w',variable= CCU1_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=CCU)
    scrollbar.insert('end', '\n')
    XGW = tk.Checkbutton(scrollbar, text='XGW', bg='white', anchor='w',variable= XGW_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=XGW)
    scrollbar.insert('end', '\n')
    APM = tk.Checkbutton(scrollbar, text='APM', bg='white', anchor='w',variable= APM_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=APM)
    scrollbar.insert('end', '\n')
    SHVU_F = tk.Checkbutton(scrollbar, text='SHVU_F', bg='white', anchor='w',variable= SHVU_F_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=SHVU_F)
    scrollbar.insert('end', '\n')
    SHVU_R = tk.Checkbutton(scrollbar, text='SHVU_R', bg='white', anchor='w',variable= SHVU_R_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=SHVU_R)
    scrollbar.insert('end', '\n')
    OCS = tk.Checkbutton(scrollbar, text='OCS', bg='white', anchor='w',variable= OCS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=OCS)
    scrollbar.insert('end', '\n')
    MHU = tk.Checkbutton(scrollbar, text='MHU', bg='white', anchor='w',variable= MHU_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MHU)
    scrollbar.insert('end', '\n')
    HUD = tk.Checkbutton(scrollbar, text='HUD', bg='white', anchor='w',variable= HUD_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=HUD)
    scrollbar.insert('end', '\n')
    AVAS = tk.Checkbutton(scrollbar, text='AVAS', bg='white', anchor='w',variable= AVAS_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=AVAS)
    scrollbar.insert('end', '\n')
    AP_ECU = tk.Checkbutton(scrollbar, text='AP_ECU', bg='white', anchor='w',variable= AP_ECU_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=AP_ECU)
    scrollbar.insert('end', '\n')
    FCAM = tk.Checkbutton(scrollbar, text='FCAM', bg='white', anchor='w',variable= FCAM_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=FCAM)
    scrollbar.insert('end', '\n')
    MCR_FL = tk.Checkbutton(scrollbar, text='MCR_FL', bg='white', anchor='w',variable= MCR_FL_RADAR_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MCR_FL)
    scrollbar.insert('end', '\n')
    MCR_FR = tk.Checkbutton(scrollbar, text='MCR_FR', bg='white', anchor='w',variable= MCR_FR_RADAR_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MCR_FR)
    scrollbar.insert('end', '\n')
    MCR_RR = tk.Checkbutton(scrollbar, text='MCR_RR', bg='white', anchor='w',variable= MCR_RR_RADAR_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MCR_RR)
    scrollbar.insert('end', '\n')
    MCR_RL = tk.Checkbutton(scrollbar, text='MCR_LL', bg='white', anchor='w',variable= MCR_RL_RADAR_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MCR_RL)
    scrollbar.insert('end', '\n')
    MFR = tk.Checkbutton(scrollbar, text='MFR', bg='white', anchor='w',variable= MFR1_RADAR_Status,onvalue=1, offvalue=0)
    scrollbar.window_create('end', window=MFR)
    scrollbar.insert('end', '\n')    

    def RUN_Read():
        Terminal_clear()
        if VCU_Status.get() == 1:
            VCU_READ_DID()           
        if DCDC_Status.get() == 1:
            POD_DCDC_READ_DID()
        if POD_Status.get() == 1:
            POD_GW_READ_DID()
        if OBC_Status.get() == 1:
            POD_OBC_READ_DID()
        if EDS_Status.get() == 1:
            EDS_F_READ_DID()
        if BMS_Status.get() == 1:
            BMS_READ_DID()
        if GS_Status.get() == 1:
            GS_READ_DID()
        if IDB_Status == 1:
            IDB_READ_DID()
        if RCU_Status.get() == 1:
            RCU_READ_DID()
        if EPS_Status.get() == 1:
            EPS_READ_DID()
        if ACM_Status.get() == 1:
            ACM_READ_DID()
        if BCM_Status.get() == 1:
            BCM_READ_DID()
        if BCM_BPM_Status.get() == 1:
            BCM_BPM_READ_DID()
        if CCU1_Status.get() == 1:
            CCUF_READ_DID()
        if XGW_Status.get() == 1:
            XGW_READ_DID()
        if APM_Status.get() == 1:
            APM_READ_DID()
        if SHVU_F_Status.get() == 1:
            SHVU_F_READ_DID()
        if SHVU_R_Status.get() == 1:
            SHVU_R_READ_DID()
        if OCS_Status.get() == 1:
            OCS_P_READ_DID()
        if MHU_Status.get() == 1:
            MHU_READ_DID()
        if HUD_Status.get() == 1:
            HUD_READ_DID()
        if AVAS_Status.get() == 1:          
            AVAS_READ_DID()
        if FCAM_Status.get() == 1:          
            PAS_READ_DID()
        if MCR_FL_RADAR_Status.get() == 1:          
            SRR_FL_READ_DID()
        if MCR_FR_RADAR_Status.get() == 1:          
            SRR_FR_READ_DID()
        if MCR_RL_RADAR_Status.get() == 1:          
            SRR_RL_READ_DID()
        if MCR_RR_RADAR_Status.get() == 1:          
            SRR_RR_READ_DID()
        if MFR1_RADAR_Status.get() == 1:          
            MRGEN_READ_DID()
# Display text for read ECU actual
    def ECU_Text(ECU,TypeRead,Header,infor):
        if Header == 1:
            text_box.insert('insert', ECU + '\n')
            text_box.update()
        else:
            text_box.insert('insert', '    ' + TypeRead + ': '+ infor + '\n')
            text_box.update()
    def Terminal_clear():
            text_box.delete(1.0,'end')
    # Write data in excel file=============================================================
# Import data from BOM VSR ECO
    def load_data_VSR_Eco():
        global ECU_LIST 
        ECU_LIST = [0 for i in range(200)] 
        Templete = load_workbook('Templete_VF6.xlsx')
        Templete_sheet = Templete.get_sheet_by_name('ECU_DID')

        count = 0
        for i in range(1,11):
            count += 1
            tree.column("# "+ str(count), anchor=CENTER)
            tree.heading("# "+ str(count), text=Templete_sheet.cell(row=1,column=i).value)
            tree.pack(expand=True, fill='y')

        # Insert the data in Treeview widget
        ID = 0
        for i in range(1,Templete_sheet.max_row+1):
            if Templete_sheet.cell(row=i,column=11).value == 'X':
                ID +=1
                ECU_LIST[ID] = str(Templete_sheet.cell(row=i,column=1).value) + str(Templete_sheet.cell(row=i,column=3).value)
                tree.insert('', 'end',iid =ECU_LIST[ID],values=(Templete_sheet.cell(row=i,column=1).value, Templete_sheet.cell(row=i,column=2).value,Templete_sheet.cell(row=i,column=3).value,"None",Templete_sheet.cell(row=i,column=5).value,"None","None","None","None",Templete_sheet.cell(row=i,column=10).value))
                tree.pack()
        
        tree.update()
        time.sleep(3)

        filetypes = (
            ('text files', '*.xlsx'),
            ('All files', '*.*')
        )
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)
        showinfo(
            title='Selected File',
            message=filename
        )
        try:
            tk.messagebox.showerror("Information", "The file you have chosen is valid")
        except ValueError:
            tk.messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            tk.messagebox.showerror("Information", f"No such file as {filename}")
            return None
        
        BOM = load_workbook(filename,data_only= True)
        BOM_sheet = BOM.worksheets[1]

        for i in range(1,BOM_sheet.max_row + 1):
            if (BOM_sheet.cell(row= i,column=13).value) == 'X':
                for j in range(1,ID):
                    if (BOM_sheet.cell(row= i,column=4).value in tree.item(ECU_LIST[j])["values"]) == True:
                        if tree.item(ECU_LIST[j])["values"][2] == BOM_sheet.cell(row= i,column=6).value:
                            if  BOM_sheet.cell(row= i,column=6).value =='F101':
                                ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                            BOM_sheet.cell(row= i,column=3).value,tree.item(ECU_LIST[j])["values"][4],tree.item(ECU_LIST[j])["values"][5],
                                            tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                                break
                            else:
                                ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                            BOM_sheet.cell(row= i,column=3).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=3).value[12:len(BOM_sheet.cell(row= i,column=3).value)],
                                            tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                                break
      
# Import data from BOM VSR PLUS
    def load_data_VSR_Plus():
        global ECU_LIST 
        ECU_LIST = [0 for i in range(200)] 
        Templete = load_workbook('Templete_VF6.xlsx')
        Templete_sheet = Templete.get_sheet_by_name('ECU_DID')

        count = 0
        for i in range(1,11):
            count += 1
            tree.column("# "+ str(count), anchor=CENTER)
            tree.heading("# "+ str(count), text=Templete_sheet.cell(row=1,column=i).value)
            tree.pack(expand=True, fill='y')

        # Insert the data in Treeview widget
        ID = 0
        for i in range(1,Templete_sheet.max_row+1):
            if Templete_sheet.cell(row=i,column=12).value == 'X':
                ID +=1
                ECU_LIST[ID] = str(Templete_sheet.cell(row=i,column=1).value) + str(Templete_sheet.cell(row=i,column=3).value)
                tree.insert('', 'end',iid =ECU_LIST[ID],values=(Templete_sheet.cell(row=i,column=1).value, Templete_sheet.cell(row=i,column=2).value,Templete_sheet.cell(row=i,column=3).value,"None",Templete_sheet.cell(row=i,column=5).value,"None","None","None","None",Templete_sheet.cell(row=i,column=10).value))
                tree.pack()
        
        tree.update()
        time.sleep(3)

        filetypes = (
            ('text files', '*.xlsx'),
            ('All files', '*.*')
        )
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)
        showinfo(
            title='Selected File',
            message=filename
        )
        try:
            tk.messagebox.showerror("Information", "The file you have chosen is valid")
        except ValueError:
            tk.messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            tk.messagebox.showerror("Information", f"No such file as {filename}")
            return None
        
        BOM = load_workbook(filename,data_only= True)
        BOM_sheet = BOM.worksheets[1]

        for i in range(1,BOM_sheet.max_row + 1):
            if (BOM_sheet.cell(row= i,column=14).value) == 'X':
                for j in range(1,ID):
                    if (BOM_sheet.cell(row= i,column=4).value in tree.item(ECU_LIST[j])["values"]) == True:
                        if tree.item(ECU_LIST[j])["values"][2] == BOM_sheet.cell(row= i,column=6).value:
                            if  BOM_sheet.cell(row= i,column=6).value =='F101':
                                ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                            BOM_sheet.cell(row= i,column=3).value,tree.item(ECU_LIST[j])["values"][4],tree.item(ECU_LIST[j])["values"][5],
                                            tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                                break
                            else:
                                ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                            BOM_sheet.cell(row= i,column=3).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=3).value[12:len(BOM_sheet.cell(row= i,column=3).value)],
                                            tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                                break

# Import Bom from JIRA ECO
    def load_data_Jira_Eco():
        global ECU_LIST 
        ECU_LIST = [0 for i in range(200)] 
        Templete = load_workbook('Templete_VF6.xlsx')
        Templete_sheet = Templete.get_sheet_by_name('ECU_DID')

        count = 0
        for i in range(1,11):
            count += 1
            tree.column("# "+ str(count), anchor=CENTER)
            tree.heading("# "+ str(count), text=Templete_sheet.cell(row=1,column=i).value)
            tree.pack(expand=True, fill='y')

        # Insert the data in Treeview widget
        ID = 0
        for i in range(1,Templete_sheet.max_row+1):
            if Templete_sheet.cell(row=i,column=11).value == 'X':
                ID +=1
                ECU_LIST[ID] = str(Templete_sheet.cell(row=i,column=1).value) + str(Templete_sheet.cell(row=i,column=3).value)
                tree.insert('', 'end',iid =ECU_LIST[ID],values=(Templete_sheet.cell(row=i,column=1).value, Templete_sheet.cell(row=i,column=2).value,Templete_sheet.cell(row=i,column=3).value,"None",Templete_sheet.cell(row=i,column=5).value,"None","None","None","None",Templete_sheet.cell(row=i,column=10).value))
                tree.pack()
        
        tree.update()
        time.sleep(3)

        filetypes = (
            ('text files', '*.xlsx'),
            ('All files', '*.*')
        )
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)
        showinfo(
            title='Selected File',
            message=filename
        )
        try:
            tk.messagebox.showerror("Information", "The file you have chosen is valid")
        except ValueError:
            tk.messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            tk.messagebox.showerror("Information", f"No such file as {filename}")
            return None
        
        BOM = load_workbook(filename,data_only= True)
        BOM_sheet = BOM.get_sheet_by_name('Rich Filter Results')  

        for i in range(1,BOM_sheet.max_row + 1):
            if ('ECO' in BOM_sheet.cell(row= i,column=6).value) == True:
                for j in range(1,ID):
                    if (BOM_sheet.cell(row= i,column=5).value in tree.item(ECU_LIST[j])["values"]) == True:
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=8).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=8).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=8).value[12:len(BOM_sheet.cell(row= i,column=8).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=9).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=9).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=9).value[12:len(BOM_sheet.cell(row= i,column=9).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=10).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=10).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=10).value[12:len(BOM_sheet.cell(row= i,column=10).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=11).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=11).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=11).value[12:len(BOM_sheet.cell(row= i,column=11).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=12).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=12).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=12).value[12:len(BOM_sheet.cell(row= i,column=12).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=13).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=13).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=13).value[12:len(BOM_sheet.cell(row= i,column=13).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=14).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=14).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=14).value[12:len(BOM_sheet.cell(row= i,column=14).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=15).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=15).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=15).value[12:len(BOM_sheet.cell(row= i,column=15).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
    
    def ImportData(ECU,PN_type,PN_DID,VF_DID,Rev_DID,VF_Rev,Real_PN_DID,Real_Rev_DID,Note,JIRA_ECU,iid):
        tree.item(iid, values=(ECU, PN_type,PN_DID,VF_DID,Rev_DID,VF_Rev,Real_PN_DID,Real_Rev_DID,Note,JIRA_ECU))
    def RealData(iid,Data,Type_Read):
        if Type_Read == 1:
            tree.item(iid, values=(tree.item(str(iid))["values"][0], tree.item(str(iid))["values"][1],tree.item(str(iid))["values"][2],tree.item(str(iid))["values"][3],tree.item(str(iid))["values"][4],tree.item(str(iid))["values"][5],Data,tree.item(str(iid))["values"][7],tree.item(str(iid))["values"][8],tree.item(str(iid))["values"][9]))
        if Type_Read == 2:
            tree.item(iid, values=(tree.item(str(iid))["values"][0], tree.item(str(iid))["values"][1],tree.item(str(iid))["values"][2],tree.item(str(iid))["values"][3],tree.item(str(iid))["values"][4],tree.item(str(iid))["values"][5],Data,tree.item(str(iid))["values"][7],tree.item(str(iid))["values"][8],tree.item(str(iid))["values"][9]))
        if Type_Read == 4:
            tree.item(iid, values=(tree.item(str(iid))["values"][0], tree.item(str(iid))["values"][1],tree.item(str(iid))["values"][2],tree.item(str(iid))["values"][3],tree.item(str(iid))["values"][4],tree.item(str(iid))["values"][5],Data,tree.item(str(iid))["values"][7],tree.item(str(iid))["values"][8],tree.item(str(iid))["values"][9]))
        elif Type_Read == 3:
            tree.item(iid, values=(tree.item(str(iid))["values"][0], tree.item(str(iid))["values"][1],tree.item(str(iid))["values"][2],tree.item(str(iid))["values"][3],tree.item(str(iid))["values"][4],tree.item(str(iid))["values"][5],tree.item(str(iid))["values"][6],Data,tree.item(str(iid))["values"][8],tree.item(str(iid))["values"][9]))
        tree.update()
# Import Bom from JIRA Plus
    def load_data_Jira_Plus():
        global ECU_LIST 
        ECU_LIST = [0 for i in range(200)] 
        Templete = load_workbook('Templete_VF6.xlsx')
        Templete_sheet = Templete.get_sheet_by_name('ECU_DID')

        count = 0
        for i in range(1,11):
            count += 1
            tree.column("# "+ str(count), anchor=CENTER)
            tree.heading("# "+ str(count), text=Templete_sheet.cell(row=1,column=i).value)
            tree.pack(expand=True, fill='y')

        # Insert the data in Treeview widget
        ID = 0
        for i in range(1,Templete_sheet.max_row+1):
            if Templete_sheet.cell(row=i,column=12).value == 'X':
                ID +=1
                tree.insert('', 'end',iid =str(ID),values=(Templete_sheet.cell(row=i,column=1).value, Templete_sheet.cell(row=i,column=2).value,Templete_sheet.cell(row=i,column=3).value,"None",Templete_sheet.cell(row=i,column=5).value,"None","None","None","None",Templete_sheet.cell(row=i,column=10).value))
                tree.pack()
        
        tree.update()
        time.sleep(3)

        filetypes = (
            ('text files', '*.xlsx'),
            ('All files', '*.*')
        )
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)
        showinfo(
            title='Selected File',
            message=filename
        )
        try:
            tk.messagebox.showerror("Information", "The file you have chosen is valid")
        except ValueError:
            tk.messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            tk.messagebox.showerror("Information", f"No such file as {filename}")
            return None
        
        BOM = load_workbook(filename,data_only= True)
        BOM_sheet = BOM.get_sheet_by_name('Rich Filter Results')  

        for i in range(1,BOM_sheet.max_row + 1):
            if ('PLUS' in BOM_sheet.cell(row= i,column=6).value) == True:
                for j in range(1,ID):
                    if (BOM_sheet.cell(row= i,column=5).value in tree.item(ECU_LIST[j])["values"]) == True:
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=8).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=8).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=8).value[12:len(BOM_sheet.cell(row= i,column=8).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=9).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=9).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=9).value[12:len(BOM_sheet.cell(row= i,column=9).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=10).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=10).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=10).value[12:len(BOM_sheet.cell(row= i,column=10).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=11).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=11).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=11).value[12:len(BOM_sheet.cell(row= i,column=11).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=12).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=12).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=12).value[12:len(BOM_sheet.cell(row= i,column=12).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=13).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=13).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=13).value[12:len(BOM_sheet.cell(row= i,column=13).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=14).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=14).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=14).value[12:len(BOM_sheet.cell(row= i,column=14).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
                            
                        if (tree.item(ECU_LIST[j])["values"][2] in BOM_sheet.cell(row= 1,column=15).value) == True:
                            ImportData(tree.item(ECU_LIST[j])["values"][0],tree.item(ECU_LIST[j])["values"][1],tree.item(ECU_LIST[j])["values"][2],
                                        BOM_sheet.cell(row= i,column=15).value[0:11],tree.item(ECU_LIST[j])["values"][4],BOM_sheet.cell(row= i,column=15).value[12:len(BOM_sheet.cell(row= i,column=15).value)],
                                        tree.item(ECU_LIST[j])["values"][6],tree.item(ECU_LIST[j])["values"][7],tree.item(ECU_LIST[j])["values"][8],tree.item(ECU_LIST[j])["values"][9],ECU_LIST[j])
#Load templete file bom
    menubar1 = Menu(VF6) 
    file = Menu(menubar1, tearoff = 0) 
    menubar1.add_cascade(label ='Import', menu = file) 
    file.add_separator()     
    file.add_command(label ='BOM VSR ECO', command = load_data_VSR_Eco)
    file.add_command(label ='BOM VSR PLUS', command = load_data_VSR_Plus)    
    file.add_command(label ='BOM JIRA ECO', command = load_data_Jira_Eco)
    file.add_command(label ='BOM JIRA PLUS', command = load_data_Jira_Plus)

# Adding Help Menu 
    RUN = Menu(menubar1, tearoff = 0) 
    menubar1.add_cascade(label ='RUN', menu = RUN) 
    RUN.add_separator() 
    RUN.add_command(label ='RUN', command = RUN_Read)    

    Export = Menu(menubar1, tearoff = 0) 
    menubar1.add_cascade(label ='Export', menu = Export) 
    RUN.add_separator() 
    RUN.add_command(label ='Export', command = RUN_Read)   

    VF6.config(menu = menubar1) 
#====================================VF7 Window============================================		
def VF7Window():     
    newWindow = Toplevel(root)
    newWindow.title("New Window")
    newWindow.geometry("200x200") 
    Label(newWindow, 
          text ="This is a new VF7").pack()
#====================================VF8 Window============================================		
def VF8Window():     
    newWindow = Toplevel(root)
    newWindow.title("New Window")
    newWindow.geometry("200x200") 
    Label(newWindow, 
          text ="This is a new VF8").pack()
#====================================VF9 Window============================================		
def VF9Window():     
    newWindow = Toplevel(root)
    newWindow.title("New Window")
    newWindow.geometry("200x200") 
    Label(newWindow, 
          text ="This is a new VF9").pack()
#================================================================================		          
# Creating Menubar 
menubar = Menu(root) 
# Adding File Menu and commands 
file = Menu(menubar, tearoff = 0) 
menubar.add_cascade(label ='File', menu = file) 
file.add_separator()
file.add_command(label ='Hardware Congfig', command = Hardware_Config) 
file.add_command(label ='Import DBC', command = OpenDBC) 
file.add_command(label ='Exit', command = root.destroy) 
  
# Adding Edit Menu and commands 
edit = Menu(menubar, tearoff = 0) 
menubar.add_cascade(label ='Model', menu = edit) 
edit.add_separator() 
edit.add_command(label ='VF3', command = VF3Window) 
edit.add_command(label ='VF5', command = VF5Window)
edit.add_command(label ='VF6', command = VF6Window) 
edit.add_command(label ='VF7', command = VF7Window) 
edit.add_command(label ='VF8', command = VF8Window) 
edit.add_command(label ='VF9', command = VF9Window) 
  
# Adding Help Menu 
help_ = Menu(menubar, tearoff = 0) 
menubar.add_cascade(label ='Help', menu = help_) 
help_.add_command(label ='Tk Help', command = None) 
help_.add_command(label ='Demo', command = None) 
help_.add_separator() 
help_.add_command(label ='About Tk', command = None)
  
# display Menu 
root.config(menu = menubar)
mainloop() 